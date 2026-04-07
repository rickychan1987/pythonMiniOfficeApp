import pandas as pd
import xlrd
from xlrd import open_workbook
import glob
import re
from openpyxl import load_workbook

def clean_text_before_marker(text, marker="#&"):
    """
    Extract text before the marker (like #& or #&VN)
    Also removes common suffixes
    """
    if text is None:
        return text
    
    if isinstance(text, str):
        # Method 1: Split by #&
        if marker in text:
            text = text.split(marker)[0]
        
        # Method 2: Also handle other possible separators
        for sep in ['#&VN', '#&', '##', '|', ';', '.']:
            if sep in text:
                text = text.split(sep)[0]
        
        # Remove trailing dots or spaces
        text = text.strip().rstrip('.')
        
    return text

def extract_excel_data_xls(file_path):
    """
    Extract specific data from .xls file
    F_rows: with cleaning (remove #&VN)
    Q_rows: WITHOUT cleaning (raw data)
    R_rows: WITHOUT cleaning (raw data)
    """
    wb = xlrd.open_workbook(file_path, formatting_info=False)
    ws = wb.sheet_by_index(0)
    
    extracted_data = {}
    
    # Extract and clean single cell data
    extracted_data['E4'] = clean_text_before_marker(ws.cell_value(3, 4))
    extracted_data['F30'] = (ws.cell_value(29, 5))
    extracted_data['F33'] = clean_text_before_marker(ws.cell_value(32, 5))
    extracted_data['R33'] = clean_text_before_marker(ws.cell_value(32, 17))
    extracted_data['R49'] = clean_text_before_marker(ws.cell_value(48, 17))
    extracted_data['U53'] = ws.cell_value(52, 20)
    
    # F-158 182 215 239 272 296 - WITH cleaning
    f_rows = [158, 182, 215, 239, 272, 296]
    extracted_data['F_rows'] = []
    for row in f_rows:
        try:
            raw_value = ws.cell_value(row-1, 5)
            cleaned_value = clean_text_before_marker(raw_value)
            extracted_data['F_rows'].append(cleaned_value)
        except IndexError:
            extracted_data['F_rows'].append(f"Row {row} out of range")
    
    # Q-161 185 218 242 275 299 - WITHOUT cleaning
    q_rows = [161, 185, 218, 242, 275, 299]
    extracted_data['Q_rows'] = []
    for row in q_rows:
        try:
            raw_value = ws.cell_value(row-1, 16)
            extracted_data['Q_rows'].append(raw_value)
        except IndexError:
            extracted_data['Q_rows'].append(f"Row {row} out of range")
    
    # R-163 187 220 244 277 301 - WITHOUT cleaning
    r_rows = [163, 187, 220, 244, 277, 301]
    extracted_data['R_rows'] = []
    for row in r_rows:
        try:
            raw_value = ws.cell_value(row-1, 17)
            extracted_data['R_rows'].append(raw_value)
        except IndexError:
            extracted_data['R_rows'].append(f"Row {row} out of range")
    
    # Store row lists
    extracted_data['f_row_list'] = f_rows
    extracted_data['q_row_list'] = q_rows
    extracted_data['r_row_list'] = r_rows
    
    return extracted_data

def save_to_single_sheet(extracted_data, output_file):
    """
    Save ALL extracted data to ONE sheet with blank rows as separators
    """
    if output_file.endswith('.xls'):
        output_file = output_file.replace('.xls', '.xlsx')
    
    # Create a list to hold all rows for the single sheet
    all_rows = []
    
    # Section 1: Single Cells
    all_rows.append(['=== SINGLE CELL VALUES ==='])
    all_rows.append(['Cell Reference', 'Value'])
    all_rows.append(['E4', extracted_data['E4']])
    all_rows.append(['F30', extracted_data['F30']])
    all_rows.append(['F33', extracted_data['F33']])
    all_rows.append(['R33', extracted_data['R33']])
    all_rows.append(['R49', extracted_data['R49']])
    all_rows.append(['U53', extracted_data['U53']])
    all_rows.append([])  # Blank row separator
    all_rows.append([])  # Blank row separator
    
    # Section 2: Column F - Mô tả hàng hóa (Cleaned)
    all_rows.append(['=== COLUMN F - MÔ TẢ HÀNG HÓA (CLEANED) ==='])
    all_rows.append(['Row', 'Mô tả hàng hóa (Cleaned)'])
    for i, row in enumerate(extracted_data['f_row_list']):
        all_rows.append([row, extracted_data['F_rows'][i]])
    all_rows.append([])  # Blank row separator
    all_rows.append([])  # Blank row separator
    
    # Section 3: Column Q - Số lượng (Raw)
    all_rows.append(['=== COLUMN Q - SỐ LƯỢNG (RAW) ==='])
    all_rows.append(['Row', 'Số lượng (Raw)'])
    for i, row in enumerate(extracted_data['q_row_list']):
        all_rows.append([row, extracted_data['Q_rows'][i]])
    all_rows.append([])  # Blank row separator
    all_rows.append([])  # Blank row separator
    
    # Section 4: Column R - Đơn giá (Raw)
    all_rows.append(['=== COLUMN R - ĐƠN GIÁ (RAW) ==='])
    all_rows.append(['Row', 'Đơn giá (Raw)'])
    for i, row in enumerate(extracted_data['r_row_list']):
        all_rows.append([row, extracted_data['R_rows'][i]])
    
    # Create DataFrame from all rows
    df_all = pd.DataFrame(all_rows)
    
    # Save to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_all.to_excel(writer, sheet_name='Extracted_Data', index=False, header=False)
    
    # Auto-adjust column widths
    workbook = load_workbook(output_file)
    worksheet = workbook['Extracted_Data']
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output_file)
    
    print(f"✅ All data saved to single sheet in: {output_file}")

def display_extracted_data(extracted_data):
    """
    Display extracted data
    """
    print("\n" + "="*70)
    print("EXTRACTED DATA SUMMARY")
    print("="*70)
    
    print("\n📌 Single Cell Values:")
    print("-" * 50)
    print(f"  E4  → {extracted_data['E4']}")
    print(f"  F30 → {extracted_data['F30']}")
    print(f"  F33 → {extracted_data['F33']}")
    print(f"  R33 → {extracted_data['R33']}")
    print(f"  R49 → {extracted_data['R49']}")
    print(f"  U53 → {extracted_data['U53']}")
    
    print("\n📌 Column F - Mô tả hàng hóa (CLEANED):")
    print("-" * 50)
    for i, row in enumerate(extracted_data['f_row_list']):
        value = str(extracted_data['F_rows'][i])[:100]
        print(f"  Row {row}: {value}...")
    
    print("\n📌 Column Q - Số lượng (RAW):")
    print("-" * 50)
    for i, row in enumerate(extracted_data['q_row_list']):
        print(f"  Row {row}: {extracted_data['Q_rows'][i]}")
    
    print("\n📌 Column R - Đơn giá (RAW):")
    print("-" * 50)
    for i, row in enumerate(extracted_data['r_row_list']):
        print(f"  Row {row}: {extracted_data['R_rows'][i]}")

def process_all_excel_files():
    """
    Process all .xls files in current directory
    """
    excel_files = glob.glob("*.xls")
    excel_files.extend(glob.glob("*.xlsx"))
    excel_files = [f for f in excel_files if not f.startswith("extracted_") and not f.startswith("result_")]
    
    if not excel_files:
        print("No Excel files found!")
        return
    
    print(f"Found {len(excel_files)} Excel file(s)")
    
    for file in excel_files:
        print(f"\n{'='*70}")
        print(f"📁 Processing: {file}")
        print('='*70)
        
        try:
            if file.endswith('.xls'):
                extracted_data = extract_excel_data_xls(file)
            else:
                # For .xlsx files
                from openpyxl import load_workbook
                wb = load_workbook(file, data_only=True)
                ws = wb.active
                
                f_rows = [158, 182, 215, 239, 272, 296]
                q_rows = [161, 185, 218, 242, 275, 299]
                r_rows = [163, 187, 220, 244, 277, 301]
                
                extracted_data = {
                    'E4': clean_text_before_marker(ws['E4'].value),
                    'F30': clean_text_before_marker(ws['F30'].value),
                    'F33': clean_text_before_marker(ws['F33'].value),
                    'R33': clean_text_before_marker(ws['R33'].value),
                    'R49': clean_text_before_marker(ws['R49'].value),
                    'U53': ws['U53'].value,
                    'F_rows': [clean_text_before_marker(ws[f'F{row}'].value) for row in f_rows],
                    'Q_rows': [ws[f'Q{row}'].value for row in q_rows],
                    'R_rows': [ws[f'R{row}'].value for row in r_rows],
                    'f_row_list': f_rows,
                    'q_row_list': q_rows,
                    'r_row_list': r_rows
                }
            
            display_extracted_data(extracted_data)
            
            # Create output filename
            base_name = file.replace('.xls', '').replace('.xlsx', '')
            output_file = f"extracted_{base_name}.xlsx"
            
            save_to_single_sheet(extracted_data, output_file)
            
        except Exception as e:
            print(f"❌ Error processing {file}: {e}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    # Test the cleaning function
    test_text = "Giá treo tivi cố định gắn tường (TV Mount), model: PIMFK3, chất liệu kim loại thép, kích thước: 487*430*418mm. Hàng mới 100%#&VN"
    cleaned = clean_text_before_marker(test_text)
    print("Testing text cleaning:")
    print(f"Original: {test_text}")
    print(f"Cleaned:  {cleaned}")
    print("\n" + "="*70)
    
    # Process all Excel files
    process_all_excel_files()