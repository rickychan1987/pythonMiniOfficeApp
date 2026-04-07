import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# --- CONFIGURATION ---
base_folder = r"D:\Administrator\Pictures\Camera Roll\tong_hop_folder"
merged_output = os.path.join(base_folder, "total_name_list.xlsx")

# --- STEP 6: Merge all *_photo_links.xlsx files ---
wb_merged = Workbook()
ws_merged = wb_merged.active
ws_merged.title = "All Photo Links"

# Write header row
ws_merged.append(["ID", "Name", "Photo Link"])
for cell in ws_merged[1]:
    cell.font = Font(bold=True)

row_counter = 2

for root, _, files in os.walk(base_folder):
    for file in files:
        if file.endswith("_photo_links.xlsx"):
            file_path = os.path.join(root, file)
            try:
                wb = load_workbook(file_path)
                ws = wb.active

                # Skip header (start from 2nd row)
                for r in ws.iter_rows(min_row=2, values_only=False):
                    id_val = r[0].value
                    name_val = r[1].value
                    photo_cell = r[2]

                    # Copy hyperlink and displayed text if exists
                    if photo_cell.hyperlink:
                        new_cell = ws_merged.cell(row=row_counter, column=3, value=photo_cell.value)
                        new_cell.hyperlink = photo_cell.hyperlink.target
                        new_cell.font = Font(color="0000FF", underline="single")
                    else:
                        ws_merged.cell(row=row_counter, column=3, value=photo_cell.value)

                    # Copy ID, Name, and Source File
                    ws_merged.cell(row=row_counter, column=1, value=id_val)
                    ws_merged.cell(row=row_counter, column=2, value=name_val)
                    ws_merged.cell(row=row_counter, column=4, value=file)

                    row_counter += 1

                wb.close()
                print(f"✅ Merged: {file}")
            except Exception as e:
                print(f"⚠️ Error reading {file}: {e}")

# Auto-adjust column widths
for col in ["A", "B", "C", "D"]:
    ws_merged.column_dimensions[col].width = 25 if col != "C" else 60

# Save merged file
wb_merged.save(merged_output)
print(f"🎯 Step 6 complete: Merged Excel with working hyperlinks → {merged_output}")
