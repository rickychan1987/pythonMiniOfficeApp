import sys
from typing import List, Dict, Any

def read_with_openpyxl(path: str, sheet: str | int = 0, header: bool = True) -> List[Dict[str, Any]] | List[List[Any]]:
    """
    Read .xlsx using openpyxl.
    - sheet: 0-based index or sheet name
    - header: if True, return list of dicts using first row as keys; otherwise return list of rows (lists).
    """
    from openpyxl import load_workbook
    wb = load_workbook(filename=path, data_only=True)
    # select sheet
    if isinstance(sheet, int):
        ws = wb[wb.sheetnames[sheet]]
    else:
        ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    if header:
        keys = [str(k) if k is not None else f"col{i}" for i, k in enumerate(rows[0], start=1)]
        return [dict(zip(keys, row)) for row in rows[1:]]
    else:
        return [list(r) for r in rows]

def read_with_pandas(path: str, sheet: str | int = 0, header_row: int | None = 0):
    """
    Read .xlsx using pandas. header_row=None returns raw dataframe without header.
    """
    import pandas as pd
    df = pd.read_excel(path, sheet_name=sheet, header=header_row)
    return df  # caller can use df.to_dict(orient='records') or df.values

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python read_xlsx.py <file.xlsx> [sheet_name_or_index]")
        sys.exit(1)
    path = sys.argv[1]
    sheet_arg = sys.argv[2] if len(sys.argv) > 2 else 0
    # try parse sheet index
    try:
        sheet = int(sheet_arg)
    except Exception:
        sheet = sheet_arg
    try:
        data = read_with_openpyxl(path, sheet=sheet, header=True)
        print("Read with openpyxl (list of dicts):")
        for r in data[:10]:
            print(r)
    except Exception as e:
        print("openpyxl read failed:", e)
        try:
            df = read_with_pandas(path, sheet=sheet, header_row=0)
            print("Read with pandas (first 10 rows):")
            print(df.head(10))
        except Exception as e2:
            print("pandas read failed:", e2)
            sys.exit(2)